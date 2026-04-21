from __future__ import annotations

import io
import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from device_inventory import DeviceInventoryRepository


def write_json(path: Path, key: str, records: list[dict]) -> None:
    path.write_text(json.dumps({key: records}, ensure_ascii=False, indent=2), encoding="utf-8")


class DeviceInventoryReportWorkbookTestCase(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        base = Path(self.temp_dir.name)
        self.inventory_path = base / "device_inventory.json"
        self.events_path = base / "device_inventory_events.json"
        write_json(
            self.inventory_path,
            "devices",
            [
                {
                    "id": "dev-1",
                    "management_no": "A-001",
                    "asset_group": "노트북",
                    "location": "컴퓨터실",
                    "device_type": "노트북",
                    "manufacturer": "삼성",
                    "model_name": "Galaxy Book",
                    "serial_no": "SN-001",
                    "cpu": "i5",
                    "ram": "16GB",
                    "introduced_date": "2020-03-01",
                    "status": "정상 사용",
                    "notes": "",
                    "user_name": "홍길동",
                    "image_url": "",
                    "created_at": "2026-04-20T00:00:00+00:00",
                    "updated_at": "2026-04-20T00:00:00+00:00",
                },
                {
                    "id": "dev-2",
                    "management_no": "A-002",
                    "asset_group": "데스크톱",
                    "location": "교무실",
                    "device_type": "데스크톱",
                    "manufacturer": "LG",
                    "model_name": "Desk Pro",
                    "serial_no": "SN-002",
                    "cpu": "i7",
                    "ram": "32GB",
                    "introduced_date": "2018-02-01",
                    "status": "점검 필요",
                    "notes": "팬 소음",
                    "user_name": "이순신",
                    "image_url": "",
                    "created_at": "2026-04-20T00:00:00+00:00",
                    "updated_at": "2026-04-20T00:00:00+00:00",
                },
                {
                    "id": "dev-3",
                    "management_no": "A-003",
                    "asset_group": "모니터",
                    "location": "",
                    "device_type": "모니터",
                    "manufacturer": "Dell",
                    "model_name": "UltraSharp",
                    "serial_no": "SN-003",
                    "cpu": "",
                    "ram": "",
                    "introduced_date": "2017-05-01",
                    "status": "정상 사용",
                    "notes": "",
                    "user_name": "",
                    "image_url": "",
                    "created_at": "2026-04-20T00:00:00+00:00",
                    "updated_at": "2026-04-20T00:00:00+00:00",
                },
            ],
        )
        write_json(self.events_path, "events", [])
        self.repository = DeviceInventoryRepository(self.inventory_path, self.events_path)

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def load_report(self):
        payload = self.repository.export_report_workbook()
        return load_workbook(io.BytesIO(payload))

    def test_report_contains_dashboard_total_sheet_and_location_tabs(self) -> None:
        workbook = self.load_report()

        self.assertEqual(workbook.sheetnames[0], "대시보드")
        self.assertEqual(workbook.sheetnames[1], "전체대장")
        self.assertIn("컴퓨터실", workbook.sheetnames)
        self.assertIn("교무실", workbook.sheetnames)
        self.assertIn("미지정", workbook.sheetnames)

    def test_dashboard_contains_title_and_location_summary(self) -> None:
        dashboard = self.load_report()["대시보드"]

        self.assertEqual(dashboard["A1"].value, "DCMS 기기관리대장")
        self.assertEqual(dashboard["A2"].value, "설치장소별 자산 분포 요약")
        self.assertEqual(dashboard["A1"].fill.fgColor.rgb[-6:], "1F4E79")
        self.assertEqual(dashboard["A1"].font.color.rgb[-6:], "FFFFFF")
        self.assertEqual(dashboard["B5"].value, 3)
        self.assertEqual(dashboard["D5"].value, 3)
        self.assertEqual(dashboard["E10"].value, "교무실")
        self.assertEqual(dashboard["F10"].value, 1)

    def test_dashboard_status_summary_uses_serialized_flags(self) -> None:
        dashboard = self.load_report()["대시보드"]

        self.assertEqual(dashboard["B8"].value, 2)
        self.assertEqual(dashboard["D8"].value, 1)
        self.assertEqual(dashboard["F8"].value, 3)

    def test_location_sheet_contains_only_matching_rows(self) -> None:
        sheet = self.load_report()["교무실"]

        self.assertEqual(sheet["A1"].value, "교무실")
        self.assertEqual(sheet["A4"].value, "관리번호")
        self.assertEqual(sheet["A5"].value, "A-002")
        self.assertEqual(sheet.max_row, 5)

    def test_full_inventory_sheet_has_filter_freeze_and_header_style(self) -> None:
        sheet = self.load_report()["전체대장"]

        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(sheet.auto_filter.ref, f"A1:O{sheet.max_row}")
        self.assertTrue(sheet["A1"].font.bold)
        self.assertEqual(sheet["A1"].fill.fgColor.rgb[-6:], "EEF2F6")
        self.assertEqual(sheet["A1"].border.bottom.color.rgb[-6:], "B8C4D3")

    def test_location_sheet_has_local_heading_and_filtered_table(self) -> None:
        sheet = self.load_report()["컴퓨터실"]

        self.assertEqual(sheet["A1"].value, "컴퓨터실")
        self.assertEqual(sheet["A2"].value, "장비 수")
        self.assertEqual(sheet["B2"].value, 1)
        self.assertEqual(sheet.freeze_panes, "A5")
        self.assertEqual(sheet["A4"].value, "관리번호")


if __name__ == "__main__":
    unittest.main()
