from __future__ import annotations

import csv
import io
import json
import threading
import uuid
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


EXPECTED_LIFE_CYCLE_YEARS = 5

BG_PAGE = "F5F7FA"
BG_SURFACE = "FFFFFF"
BG_SUBTLE = "EEF2F6"
LINE_DEFAULT = "D6DEE8"
LINE_STRONG = "B8C4D3"
TEXT_PRIMARY = "15202B"
TEXT_SECONDARY = "445266"
TEXT_MUTED = "69788C"
TEXT_INVERSE = "FFFFFF"
ACCENT_PRIMARY = "1F4E79"
ACCENT_SOFT = "DCE8F5"
SUCCESS = "2E7D32"
SUCCESS_SOFT = "E6F4EA"
WARNING = "B26A00"
WARNING_SOFT = "FFF4DD"
DANGER = "C62828"
DANGER_SOFT = "FDECEC"


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _today() -> date:
    return date.today()


class DeviceInventoryRepository:
    csv_headers = [
        "관리번호",
        "분류",
        "설치장소",
        "형태",
        "제조회사",
        "모델명",
        "시리얼넘버",
        "CPU",
        "RAM",
        "구입시기",
        "사용연수",
        "상태",
        "비고",
        "사용자명",
        "제품이미지",
    ]

    _editable_fields = [
        "management_no",
        "asset_group",
        "location",
        "device_type",
        "manufacturer",
        "model_name",
        "serial_no",
        "cpu",
        "ram",
        "introduced_date",
        "status",
        "notes",
        "user_name",
        "image_url",
    ]

    _field_aliases = {
        "관리번호": "management_no",
        "management_no": "management_no",
        "분류": "asset_group",
        "asset_group": "asset_group",
        "카테고리": "asset_group",
        "설치장소": "location",
        "location": "location",
        "형태": "device_type",
        "device_type": "device_type",
        "제조회사": "manufacturer",
        "manufacturer": "manufacturer",
        "모델명": "model_name",
        "model_name": "model_name",
        "시리얼넘버": "serial_no",
        "시리얼번호": "serial_no",
        "serial_no": "serial_no",
        "CPU": "cpu",
        "cpu": "cpu",
        "RAM": "ram",
        "ram": "ram",
        "구입시기": "introduced_date",
        "도입일자": "introduced_date",
        "introduced_date": "introduced_date",
        "상태": "status",
        "status": "status",
        "비고": "notes",
        "notes": "notes",
        "사용자명": "user_name",
        "사용자": "user_name",
        "user_name": "user_name",
        "제품이미지": "image_url",
        "이미지 URL": "image_url",
        "image_url": "image_url",
        "사용연수": "usage_years",
    }

    def __init__(self, inventory_path: Path, events_path: Path) -> None:
        self.inventory_path = inventory_path
        self.events_path = events_path
        self._lock = threading.Lock()
        self._ensure_files()

    def list_devices(self, filters: dict | None = None) -> list[dict]:
        filters = filters or {}
        with self._lock:
            devices = self._read_records(self.inventory_path, key="devices")

        serialized = [self._serialize_device(item) for item in self._sorted_devices(devices)]
        return [item for item in serialized if self._matches_filters(item, filters)]

    def get_device(self, identifier: str) -> dict:
        with self._lock:
            devices = self._read_records(self.inventory_path, key="devices")
        record = self._find_device(devices, identifier)
        if record is None:
            raise KeyError(identifier)
        return self._serialize_device(record)

    def upsert_device(self, payload: dict, event_type: str = "create") -> tuple[dict, bool]:
        normalized = self._build_device_record(payload)
        with self._lock:
            devices = self._read_records(self.inventory_path, key="devices")
            existing_index = self._find_device_index(devices, normalized["management_no"])
            now = utc_now_iso()

            if existing_index is None:
                devices.append(normalized)
                self._write_records(self.inventory_path, key="devices", records=devices)
                self._append_event(
                    device_id=normalized["id"],
                    management_no=normalized["management_no"],
                    event_type=event_type,
                    event_summary=f'{normalized["management_no"]} 등록',
                )
                return self._serialize_device(normalized), True

            existing = devices[existing_index]
            updated = self._merge_device_record(existing, normalized, keep_identity=True)
            updated["created_at"] = existing.get("created_at", now)
            updated["updated_at"] = now
            devices[existing_index] = updated
            self._write_records(self.inventory_path, key="devices", records=devices)
            self._append_event(
                device_id=updated["id"],
                management_no=updated["management_no"],
                event_type="import" if event_type == "import" else "update",
                event_summary=f'{updated["management_no"]} 갱신',
            )
            return self._serialize_device(updated), False

    def update_device(self, identifier: str, payload: dict) -> dict:
        with self._lock:
            devices = self._read_records(self.inventory_path, key="devices")
            index = self._find_device_index(devices, identifier)
            if index is None:
                raise KeyError(identifier)

            existing = devices[index]
            updated = self._merge_device_record(existing, payload, keep_identity=False)
            updated["id"] = existing["id"]
            updated["created_at"] = existing.get("created_at", utc_now_iso())
            updated["updated_at"] = utc_now_iso()
            devices[index] = updated
            self._write_records(self.inventory_path, key="devices", records=devices)
            self._append_event(
                device_id=updated["id"],
                management_no=updated["management_no"],
                event_type="update",
                event_summary=f'{updated["management_no"]} 수정',
            )
            return self._serialize_device(updated)

    def delete_device(self, identifier: str) -> dict:
        with self._lock:
            devices = self._read_records(self.inventory_path, key="devices")
            index = self._find_device_index(devices, identifier)
            if index is None:
                raise KeyError(identifier)

            deleted = devices.pop(index)
            self._write_records(self.inventory_path, key="devices", records=devices)
            self._append_event(
                device_id=deleted["id"],
                management_no=deleted["management_no"],
                event_type="delete",
                event_summary=f'{deleted["management_no"]} 삭제',
            )
            return self._serialize_device(deleted)

    def import_csv(self, csv_text: str) -> dict:
        text = self._normalize_text(csv_text).lstrip("\ufeff")
        if not text:
            raise ValueError("CSV 내용이 비어 있습니다.")

        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise ValueError("CSV 헤더를 읽을 수 없습니다.")

        created_count = 0
        updated_count = 0
        processed_items: list[dict] = []

        for row in reader:
            if not row:
                continue
            payload = self._payload_from_csv_row(row)
            if not any(payload.values()):
                continue
            device, created = self.upsert_device(payload, event_type="import")
            processed_items.append(device)
            if created:
                created_count += 1
            else:
                updated_count += 1

        return {
            "row_count": len(processed_items),
            "created": created_count,
            "updated": updated_count,
            "upserted": len(processed_items),
            "items": processed_items,
        }

    def export_csv(self) -> str:
        devices = self.list_devices()
        buffer = io.StringIO()
        writer = csv.DictWriter(buffer, fieldnames=self.csv_headers)
        writer.writeheader()
        for device in devices:
            writer.writerow(self._device_to_csv_row(device))
        return "\ufeff" + buffer.getvalue()

    def export_report_workbook(self) -> bytes:
        workbook = Workbook()
        devices = self.list_devices()
        workbook.active.title = "대시보드"

        self._write_dashboard_sheet(workbook["대시보드"], devices)
        self._write_full_inventory_sheet(workbook.create_sheet("전체대장"), devices)

        used_titles = {"대시보드", "전체대장"}
        for location_title, location_devices in self._group_devices_by_location(devices):
            safe_title = self._normalize_sheet_title(location_title, used_titles)
            used_titles.add(safe_title)
            self._write_location_inventory_sheet(workbook.create_sheet(safe_title), safe_title, location_devices)

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def list_events(self, limit: int = 50, event_type: str | None = None) -> list[dict]:
        safe_limit = max(1, min(limit, 100))
        with self._lock:
            events = self._read_records(self.events_path, key="events")

        if event_type:
            events = [item for item in events if item.get("event_type") == event_type]
        return events[:safe_limit]

    def summarize_devices(self, filters: dict | None = None) -> dict:
        devices = self.list_devices(filters)
        return {
            "total": len(devices),
            "normal_use": sum(1 for item in devices if item.get("status") == "정상 사용"),
            "life_cycle_due": sum(1 for item in devices if item.get("life_cycle_due")),
            "repair_or_inspection_needed": sum(1 for item in devices if item.get("repair_or_inspection_needed")),
        }

    def _group_devices_by_location(self, devices: list[dict]) -> list[tuple[str, list[dict]]]:
        grouped: dict[str, list[dict]] = {}
        for device in devices:
            location = self._normalize_sheet_title(device.get("location", ""))
            grouped.setdefault(location, []).append(device)
        return sorted(grouped.items(), key=lambda item: item[0].casefold())

    def _normalize_sheet_title(self, raw_value: object, existing_titles: set[str] | None = None) -> str:
        base = self._normalize_text(raw_value) or "미지정"
        invalid_chars = set("[]:*?/\\")
        cleaned = "".join(char for char in base if char not in invalid_chars).strip() or "미지정"
        cleaned = cleaned[:31].rstrip() or "미지정"
        existing_titles = existing_titles or set()
        if cleaned not in existing_titles:
            return cleaned

        counter = 2
        while True:
            suffix = f" ({counter})"
            trimmed = cleaned[: 31 - len(suffix)].rstrip()
            candidate = f"{trimmed}{suffix}"
            if candidate not in existing_titles:
                return candidate
            counter += 1

    def _summarize_locations(self, devices: list[dict]) -> list[tuple[str, int]]:
        counts: dict[str, int] = {}
        for device in devices:
            location = self._normalize_text(device.get("location")) or "미지정"
            counts[location] = counts.get(location, 0) + 1
        return sorted(counts.items(), key=lambda item: (-item[1], item[0].casefold()))

    def _summarize_statuses(self, devices: list[dict]) -> dict[str, int]:
        return {
            "정상 사용": sum(1 for item in devices if item.get("status") == "정상 사용"),
            "점검 필요": sum(1 for item in devices if item.get("repair_or_inspection_needed")),
            "교체 검토": sum(1 for item in devices if item.get("life_cycle_due")),
        }

    def _write_dashboard_sheet(self, sheet, devices: list[dict]) -> None:
        title_fill = PatternFill(fill_type="solid", fgColor=ACCENT_PRIMARY)
        header_fill = PatternFill(fill_type="solid", fgColor=BG_SUBTLE)
        subtle_fill = PatternFill(fill_type="solid", fgColor=BG_PAGE)
        accent_fill = PatternFill(fill_type="solid", fgColor=ACCENT_SOFT)
        surface_fill = PatternFill(fill_type="solid", fgColor=BG_SURFACE)
        border = self._build_thin_border()
        strong_border = self._build_strong_border()
        location_rows = self._summarize_locations(devices)
        status_summary = self._summarize_statuses(devices)

        sheet.merge_cells("A1:F1")
        sheet.merge_cells("A2:F2")
        sheet["A1"] = "DCMS 기기관리대장"
        sheet["A2"] = "설치장소별 자산 분포 요약"
        sheet["A1"].font = Font(size=18, bold=True, color=TEXT_INVERSE)
        sheet["A2"].font = Font(size=11, color=TEXT_SECONDARY)
        sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
        sheet["A2"].alignment = Alignment(horizontal="left", vertical="center")
        sheet["A1"].fill = title_fill
        sheet["A2"].fill = header_fill
        sheet["A1"].border = strong_border
        sheet["A2"].border = border

        sheet["A5"] = "전체 장비 수"
        sheet["B5"] = len(devices)
        sheet["C5"] = "설치장소 수"
        sheet["D5"] = len(location_rows)
        top_location = location_rows[0][0] if location_rows else "-"
        sheet["E5"] = "최다 보유 장소"
        sheet["F5"] = top_location

        sheet["A8"] = "정상 사용"
        sheet["B8"] = status_summary["정상 사용"]
        sheet["C8"] = "점검 필요"
        sheet["D8"] = status_summary["점검 필요"]
        sheet["E8"] = "교체 검토"
        sheet["F8"] = status_summary["교체 검토"]
        sheet["E9"] = "설치장소"
        sheet["F9"] = "수량"

        for cell_ref in ["A5", "C5", "E5", "A8", "C8", "E8", "E9", "F9"]:
            sheet[cell_ref].font = Font(bold=True, color=TEXT_PRIMARY)
            sheet[cell_ref].fill = header_fill if cell_ref in {"E9", "F9"} else accent_fill
            sheet[cell_ref].border = strong_border if cell_ref in {"E9", "F9"} else border
            sheet[cell_ref].alignment = Alignment(horizontal="center", vertical="center")

        for cell_ref in ["B5", "D5", "F5", "B8", "D8", "F8"]:
            sheet[cell_ref].font = Font(size=14, bold=True, color=TEXT_PRIMARY)
            sheet[cell_ref].fill = surface_fill
            sheet[cell_ref].border = border
            sheet[cell_ref].alignment = Alignment(horizontal="center", vertical="center")

        sheet["B8"].fill = PatternFill(fill_type="solid", fgColor=SUCCESS_SOFT)
        sheet["B8"].font = Font(size=14, bold=True, color=SUCCESS)
        sheet["D8"].fill = PatternFill(fill_type="solid", fgColor=WARNING_SOFT)
        sheet["D8"].font = Font(size=14, bold=True, color=WARNING)
        sheet["F8"].fill = PatternFill(fill_type="solid", fgColor=DANGER_SOFT)
        sheet["F8"].font = Font(size=14, bold=True, color=DANGER)

        for row_index, (location, count) in enumerate(location_rows, start=10):
            location_cell = sheet.cell(row=row_index, column=5, value=location)
            count_cell = sheet.cell(row=row_index, column=6, value=count)
            for cell in [location_cell, count_cell]:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = surface_fill
                cell.font = Font(color=TEXT_PRIMARY)

        for column_letter, width in {
            "A": 16,
            "B": 12,
            "C": 16,
            "D": 12,
            "E": 18,
            "F": 14,
            "H": 4,
            "I": 12,
            "J": 12,
            "K": 12,
            "L": 12,
        }.items():
            sheet.column_dimensions[column_letter].width = width

        sheet.row_dimensions[1].height = 28
        sheet.row_dimensions[2].height = 22

        if location_rows:
            chart = BarChart()
            chart.type = "bar"
            chart.style = 2
            chart.title = "설치장소별 장비 수"
            chart.y_axis.title = "설치장소"
            chart.x_axis.title = "수량"
            chart.legend = None
            data = Reference(sheet, min_col=6, min_row=9, max_row=9 + len(location_rows))
            labels = Reference(sheet, min_col=5, min_row=10, max_row=9 + len(location_rows))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            chart.varyColors = False
            if chart.series:
                chart.series[0].graphicalProperties.solidFill = ACCENT_PRIMARY
                chart.series[0].graphicalProperties.line.solidFill = ACCENT_PRIMARY
            chart.height = 7
            chart.width = 11
            sheet.add_chart(chart, "H9")

    def _write_full_inventory_sheet(self, sheet, devices: list[dict]) -> None:
        self._write_inventory_table(sheet, devices, start_row=1)
        sheet.freeze_panes = "A2"
        self._apply_inventory_column_widths(sheet)

    def _write_location_inventory_sheet(self, sheet, title: str, devices: list[dict]) -> None:
        title_fill = PatternFill(fill_type="solid", fgColor=ACCENT_SOFT)
        meta_fill = PatternFill(fill_type="solid", fgColor=BG_SUBTLE)
        border = self._build_thin_border()
        strong_border = self._build_strong_border()

        sheet["A1"] = title
        sheet["A2"] = "장비 수"
        sheet["B2"] = len(devices)
        sheet["D2"] = "생성일"
        sheet["E2"] = _today().isoformat()

        sheet["A1"].font = Font(size=16, bold=True, color=ACCENT_PRIMARY)
        sheet["A1"].fill = title_fill
        sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
        sheet["A1"].border = strong_border

        for cell_ref in ["A2", "B2", "D2", "E2"]:
            sheet[cell_ref].border = border
            sheet[cell_ref].alignment = Alignment(horizontal="center", vertical="center")
            sheet[cell_ref].fill = PatternFill(fill_type="solid", fgColor=BG_SURFACE)
        for cell_ref in ["A2", "D2"]:
            sheet[cell_ref].font = Font(bold=True, color=TEXT_PRIMARY)
            sheet[cell_ref].fill = meta_fill

        self._write_inventory_table(sheet, devices, start_row=4)
        sheet.freeze_panes = "A5"
        self._apply_inventory_column_widths(sheet)

    def _write_inventory_table(self, sheet, devices: list[dict], *, start_row: int = 1) -> None:
        headers = self.csv_headers
        for column_index, title in enumerate(headers, start=1):
            sheet.cell(row=start_row, column=column_index, value=title)

        for row_offset, device in enumerate(devices, start=1):
            row = [self._device_to_csv_row(device).get(header, "") for header in headers]
            for column_index, value in enumerate(row, start=1):
                sheet.cell(row=start_row + row_offset, column=column_index, value=value)

        end_row = start_row + max(len(devices), 0)
        self._apply_inventory_table_style(sheet, start_row=start_row, end_row=end_row)

    def _apply_inventory_table_style(self, sheet, *, start_row: int, end_row: int) -> None:
        header_fill = PatternFill(fill_type="solid", fgColor=BG_SUBTLE)
        zebra_fill = PatternFill(fill_type="solid", fgColor=BG_PAGE)
        thin_border = self._build_thin_border()
        strong_border = self._build_strong_border()
        usage_years_col = self.csv_headers.index("사용연수") + 1
        status_col = self.csv_headers.index("상태") + 1

        for column_index in range(1, len(self.csv_headers) + 1):
            cell = sheet.cell(row=start_row, column=column_index)
            cell.font = Font(bold=True, color=TEXT_PRIMARY)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = strong_border

        for row_index in range(start_row + 1, end_row + 1):
            for column_index in range(1, len(self.csv_headers) + 1):
                cell = sheet.cell(row=row_index, column=column_index)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border
                cell.font = Font(color=TEXT_PRIMARY)
                if (row_index - start_row) % 2 == 0:
                    cell.fill = zebra_fill

            usage_cell = sheet.cell(row=row_index, column=usage_years_col)
            try:
                usage_years = int(usage_cell.value)
            except (TypeError, ValueError):
                usage_years = None
            if usage_years is not None and usage_years >= EXPECTED_LIFE_CYCLE_YEARS:
                usage_cell.fill = PatternFill(fill_type="solid", fgColor=WARNING_SOFT)
                usage_cell.font = Font(bold=True, color=WARNING)

            status_cell = sheet.cell(row=row_index, column=status_col)
            status_value = str(status_cell.value or "").strip()
            if status_value == "정상 사용":
                status_cell.fill = PatternFill(fill_type="solid", fgColor=SUCCESS_SOFT)
                status_cell.font = Font(bold=True, color=SUCCESS)
            elif status_value in {"수리중", "점검 필요", "점검중"}:
                status_cell.fill = PatternFill(fill_type="solid", fgColor=WARNING_SOFT)
                status_cell.font = Font(bold=True, color=WARNING)
            elif status_value:
                status_cell.fill = PatternFill(fill_type="solid", fgColor=DANGER_SOFT)
                status_cell.font = Font(bold=True, color=DANGER)

        sheet.auto_filter.ref = f"A{start_row}:{get_column_letter(len(self.csv_headers))}{end_row}"

    def _apply_inventory_column_widths(self, sheet) -> None:
        widths = {
            "A": 14,
            "B": 12,
            "C": 18,
            "D": 12,
            "E": 14,
            "F": 22,
            "G": 18,
            "H": 12,
            "I": 10,
            "J": 12,
            "K": 10,
            "L": 12,
            "M": 28,
            "N": 14,
            "O": 18,
        }
        for column_letter, width in widths.items():
            sheet.column_dimensions[column_letter].width = width

    def _build_thin_border(self) -> Border:
        return Border(
            left=Side(style="thin", color=LINE_DEFAULT),
            right=Side(style="thin", color=LINE_DEFAULT),
            top=Side(style="thin", color=LINE_DEFAULT),
            bottom=Side(style="thin", color=LINE_DEFAULT),
        )

    def _build_strong_border(self) -> Border:
        return Border(
            left=Side(style="thin", color=LINE_STRONG),
            right=Side(style="thin", color=LINE_STRONG),
            top=Side(style="thin", color=LINE_STRONG),
            bottom=Side(style="thin", color=LINE_STRONG),
        )

    def _build_device_record(self, payload: dict) -> dict:
        management_no = self._normalize_text(payload.get("management_no"))
        if not management_no:
            raise ValueError("관리번호는 필수입니다.")

        now = utc_now_iso()
        return {
            "id": uuid.uuid4().hex,
            "management_no": management_no,
            "asset_group": self._normalize_text(payload.get("asset_group")),
            "location": self._normalize_text(payload.get("location")),
            "device_type": self._normalize_text(payload.get("device_type")),
            "manufacturer": self._normalize_text(payload.get("manufacturer")),
            "model_name": self._normalize_text(payload.get("model_name")),
            "serial_no": self._normalize_text(payload.get("serial_no")),
            "cpu": self._normalize_text(payload.get("cpu")),
            "ram": self._normalize_text(payload.get("ram")),
            "introduced_date": self._normalize_date(payload.get("introduced_date")),
            "status": self._normalize_text(payload.get("status")) or "정상 사용",
            "notes": self._normalize_text(payload.get("notes")),
            "user_name": self._normalize_text(payload.get("user_name")),
            "image_url": self._normalize_text(payload.get("image_url")),
            "created_at": now,
            "updated_at": now,
        }

    def _merge_device_record(self, existing: dict, payload: dict, *, keep_identity: bool) -> dict:
        merged = dict(existing)
        source = payload if keep_identity else payload

        for field in self._editable_fields:
            if field not in source:
                continue
            value = source.get(field)
            if field == "introduced_date":
                merged[field] = self._normalize_date(value)
            else:
                merged[field] = self._normalize_text(value)

        if "management_no" in source:
            management_no = self._normalize_text(source.get("management_no"))
            if not management_no:
                raise ValueError("관리번호는 필수입니다.")
            merged["management_no"] = management_no

        if "status" in source and not merged.get("status"):
            merged["status"] = "정상 사용"

        return merged

    def _serialize_device(self, device: dict) -> dict:
        introduced_date = device.get("introduced_date", "")
        usage_years = self._calculate_usage_years(introduced_date)
        status = device.get("status", "")
        return {
            "id": device["id"],
            "management_no": device.get("management_no", ""),
            "asset_group": device.get("asset_group", ""),
            "location": device.get("location", ""),
            "device_type": device.get("device_type", ""),
            "manufacturer": device.get("manufacturer", ""),
            "model_name": device.get("model_name", ""),
            "serial_no": device.get("serial_no", ""),
            "cpu": device.get("cpu", ""),
            "ram": device.get("ram", ""),
            "introduced_date": introduced_date,
            "status": status,
            "notes": device.get("notes", ""),
            "user_name": device.get("user_name", ""),
            "image_url": device.get("image_url", ""),
            "created_at": device.get("created_at", ""),
            "updated_at": device.get("updated_at", ""),
            "usage_years": usage_years,
            "life_cycle_due": usage_years is not None and usage_years >= EXPECTED_LIFE_CYCLE_YEARS,
            "repair_or_inspection_needed": status in {"수리중", "점검 필요", "점검중"},
        }

    def _device_to_csv_row(self, device: dict) -> dict:
        return {
            "관리번호": device.get("management_no", ""),
            "분류": device.get("asset_group", ""),
            "설치장소": device.get("location", ""),
            "형태": device.get("device_type", ""),
            "제조회사": device.get("manufacturer", ""),
            "모델명": device.get("model_name", ""),
            "시리얼넘버": device.get("serial_no", ""),
            "CPU": device.get("cpu", ""),
            "RAM": device.get("ram", ""),
            "구입시기": device.get("introduced_date", ""),
            "사용연수": self._serialize_device(device).get("usage_years", ""),
            "상태": device.get("status", ""),
            "비고": device.get("notes", ""),
            "사용자명": device.get("user_name", ""),
            "제품이미지": device.get("image_url", ""),
        }

    def _payload_from_csv_row(self, row: dict) -> dict:
        payload: dict[str, str] = {}
        for raw_key, raw_value in row.items():
            if raw_key is None:
                continue
            key = self._field_aliases.get(self._normalize_text(raw_key))
            if not key or key == "usage_years":
                continue
            payload[key] = self._normalize_text(raw_value)
        return payload

    def _append_event(self, device_id: str, management_no: str, event_type: str, event_summary: str) -> None:
        events = self._read_records(self.events_path, key="events")
        event = {
            "event_id": uuid.uuid4().hex,
            "device_id": device_id,
            "management_no": management_no,
            "event_type": event_type,
            "event_summary": event_summary,
            "event_at": utc_now_iso(),
        }
        events.insert(0, event)
        self._write_records(self.events_path, key="events", records=events)

    def _matches_filters(self, device: dict, filters: dict) -> bool:
        query = self._normalize_text(filters.get("q") or filters.get("query"))
        if query:
            haystack = " ".join(
                [
                    device.get("management_no", ""),
                    device.get("asset_group", ""),
                    device.get("location", ""),
                    device.get("device_type", ""),
                    device.get("manufacturer", ""),
                    device.get("model_name", ""),
                    device.get("serial_no", ""),
                    device.get("cpu", ""),
                    device.get("ram", ""),
                    device.get("status", ""),
                    device.get("notes", ""),
                    device.get("user_name", ""),
                ]
            ).casefold()
            if query.casefold() not in haystack:
                return False

        for key, field in [
            ("management_no", "management_no"),
            ("asset_group", "asset_group"),
            ("device_type", "device_type"),
            ("status", "status"),
        ]:
            filter_value = self._normalize_text(filters.get(key))
            if filter_value and device.get(field) != filter_value:
                return False

        if "life_cycle_due" in filters:
            expected = self._as_bool(filters.get("life_cycle_due"))
            if expected is not None and device.get("life_cycle_due") is not expected:
                return False

        if "repair_or_inspection_needed" in filters:
            expected = self._as_bool(filters.get("repair_or_inspection_needed"))
            if expected is not None and device.get("repair_or_inspection_needed") is not expected:
                return False

        return True

    def _find_device(self, devices: list[dict], identifier: str) -> dict | None:
        index = self._find_device_index(devices, identifier)
        if index is None:
            return None
        return devices[index]

    def _find_device_index(self, devices: list[dict], identifier: str) -> int | None:
        search_value = self._normalize_text(identifier)
        for index, device in enumerate(devices):
            if device.get("id") == search_value or device.get("management_no") == search_value:
                return index
        return None

    def _sorted_devices(self, devices: list[dict]) -> list[dict]:
        return sorted(
            devices,
            key=lambda item: (
                item.get("management_no", "").casefold(),
                item.get("asset_group", "").casefold(),
                item.get("created_at", ""),
            ),
        )

    def _calculate_usage_years(self, introduced_date: str) -> int | None:
        if not introduced_date:
            return None
        try:
            start = date.fromisoformat(introduced_date)
        except ValueError:
            return None

        current = _today()
        years = current.year - start.year
        if (current.month, current.day) < (start.month, start.day):
            years -= 1
        return max(years, 0)

    def _normalize_date(self, value: object) -> str:
        raw = self._normalize_text(value)
        if not raw:
            return ""

        raw = raw.split("T", 1)[0].strip()
        for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
            try:
                return datetime.strptime(raw, pattern).date().isoformat()
            except ValueError:
                continue
        try:
            return date.fromisoformat(raw).isoformat()
        except ValueError as exc:
            raise ValueError(f"날짜 형식이 올바르지 않습니다: {raw}") from exc

    def _normalize_text(self, value: object) -> str:
        return str(value or "").strip()

    def _as_bool(self, value: object) -> bool | None:
        if value is None:
            return None
        if isinstance(value, bool):
            return value
        text = self._normalize_text(value).casefold()
        if not text:
            return None
        if text in {"1", "true", "yes", "y", "on", "예", "참"}:
            return True
        if text in {"0", "false", "no", "n", "off", "아니오", "거짓"}:
            return False
        return None

    def _ensure_files(self) -> None:
        self.inventory_path.parent.mkdir(parents=True, exist_ok=True)
        self.events_path.parent.mkdir(parents=True, exist_ok=True)

        if not self.inventory_path.exists():
            self._write_records(self.inventory_path, key="devices", records=[])
        if not self.events_path.exists():
            self._write_records(self.events_path, key="events", records=[])

    def _read_records(self, path: Path, key: str) -> list[dict]:
        payload = json.loads(path.read_text(encoding="utf-8"))
        records = payload.get(key, [])
        if not isinstance(records, list):
            raise ValueError(f"{path.name} 파일 구조가 올바르지 않습니다.")
        return records

    def _write_records(self, path: Path, key: str, records: list[dict]) -> None:
        payload = {key: records}
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
